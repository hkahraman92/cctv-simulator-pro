import csv
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Any, Optional
from tkinter import messagebox
from .config import TURKISH_TRANSLATION
from .calculations import mode_label

try:
    from PIL import ImageGrab
except ImportError:
    ImageGrab = None

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    openpyxl = None


def export_csv(path: str, last_all_results: Dict[str, List[Any]]):
    with open(path, "w", encoding="utf-8-sig", newline="") as file:
        writer = csv.writer(file, delimiter=";")
        writer.writerow(["Kamera", "Lens", "Görev", "Sınıf", "PPM", "Optik Menzil", "Etkin Menzil", "Durum"])
        for results in last_all_results.values():
            for result in results:
                for row in sorted(result.rows, key=lambda item: item.ppm, reverse=True):
                    writer.writerow(
                        [
                            row.camera,
                            row.mode,
                            row.level,
                            row.level_type,
                            f"{row.ppm:g}",
                            f"{row.optical_distance_m:.2f}",
                            "" if row.ground_distance_m <= 0 else f"{row.ground_distance_m:.2f}",
                            row.status,
                        ]
                    )


def export_png(path: str, root_win: Any, canvas_widget: Any):
    root_win.update_idletasks()
    if ImageGrab is None:
        eps_path = str(Path(path).with_suffix(".eps"))
        canvas_widget.postscript(file=eps_path, colormode="color")
        messagebox.showwarning(
            "PNG için Pillow gerekli",
            f"Pillow bulunamadığı için PNG üretilemedi. Çizim EPS olarak kaydedildi:\n{eps_path}",
        )
        return
    x = canvas_widget.winfo_rootx()
    y = canvas_widget.winfo_rooty()
    w = canvas_widget.winfo_width()
    h = canvas_widget.winfo_height()
    image = ImageGrab.grab(bbox=(x, y, x + w, y + h))
    image.save(path)


def pdf_hex_string(text: str) -> bytes:
    encoded = str(text).encode("utf-16-be")
    hex_body = encoded.hex().upper()
    return b"<FEFF" + hex_body.encode("ascii") + b">"


def write_simple_pdf(path: str, title: str, lines: List[str]):
    all_lines = [str(line) for line in lines]
    pages = [all_lines[i : i + 46] for i in range(0, len(all_lines), 46)] or [[]]

    objects = {
        1: b"<< /Type /Catalog /Pages 2 0 R >>",
        3: b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica /Encoding /WinAnsiEncoding >>",
    }
    kids = []
    for page_index, page_lines in enumerate(pages):
        page_id = 4 + page_index * 2
        content_id = page_id + 1
        kids.append(f"{page_id} 0 R")
        content_parts = [b"BT", b"/F1 11 Tf", b"50 800 Td", b"14 TL"]
        content_parts.append(pdf_hex_string(title) + b" Tj")
        content_parts.append(b"T*")
        content_parts.append(b"T*")
        for line in page_lines:
            content_parts.append(pdf_hex_string(line) + b" Tj")
            content_parts.append(b"T*")
        content_parts.append(b"ET")
        stream = b"\n".join(content_parts)
        objects[page_id] = (
            f"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 595 842] "
            f"/Resources << /Font << /F1 3 0 R >> >> /Contents {content_id} 0 R >>"
        ).encode("latin-1")
        objects[content_id] = (
            b"<< /Length " + str(len(stream)).encode("ascii") + b" >>\nstream\n"
            + stream + b"\nendstream"
        )
    objects[2] = f"<< /Type /Pages /Kids [{' '.join(kids)}] /Count {len(pages)} >>".encode("latin-1")

    max_id = max(objects)
    offsets = [0] * (max_id + 1)
    with open(path, "wb") as file:
        file.write(b"%PDF-1.4\n")
        for obj_id in range(1, max_id + 1):
            offsets[obj_id] = file.tell()
            file.write(f"{obj_id} 0 obj\n".encode("ascii"))
            file.write(objects[obj_id])
            file.write(b"\nendobj\n")
        xref_at = file.tell()
        file.write(f"xref\n0 {max_id + 1}\n".encode("ascii"))
        file.write(b"0000000000 65535 f \n")
        for obj_id in range(1, max_id + 1):
            file.write(f"{offsets[obj_id]:010d} 0000 n \n".encode("ascii"))
        file.write(
            f"trailer\n<< /Size {max_id + 1} /Root 1 0 R >>\nstartxref\n{xref_at}\n%%EOF".encode("ascii")
        )


# ---------------------------------------------------------------------------
# ASELSAN KURUMSAL PDF RAPORLAMA MOTORU (REPORTLAB ENTEGRASYONU)
# ---------------------------------------------------------------------------
_REPORTLAB_AVAILABLE = False
try:
    from reportlab.lib import colors as rl_colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import cm, mm
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, KeepTogether, HRFlowable
    )
    from reportlab.pdfgen import canvas as rl_canvas
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    _REPORTLAB_AVAILABLE = True
except ImportError:
    pass

_FONTS_INITIALIZED = False
_FONT_FAMILY = "Helvetica"
_FONT_FAMILY_BOLD = "Helvetica-Bold"
_FONT_FAMILY_ITALIC = "Helvetica-Oblique"


def _ensure_reportlab_fonts():
    global _FONTS_INITIALIZED, _FONT_FAMILY, _FONT_FAMILY_BOLD, _FONT_FAMILY_ITALIC
    if _FONTS_INITIALIZED or not _REPORTLAB_AVAILABLE:
        return
    # Windows Segoe UI font registration for 100% Turkish character support
    font_candidates = [
        ("SegoeUI", "C:/Windows/Fonts/segoeui.ttf", "SegoeUI-Bold", "C:/Windows/Fonts/segoeuib.ttf", "SegoeUI-Italic", "C:/Windows/Fonts/segoeuii.ttf"),
        ("Arial", "C:/Windows/Fonts/arial.ttf", "Arial-Bold", "C:/Windows/Fonts/arialbd.ttf", "Arial-Italic", "C:/Windows/Fonts/ariali.ttf"),
    ]
    for reg_name, reg_path, bold_name, bold_path, ital_name, ital_path in font_candidates:
        try:
            if Path(reg_path).exists() and Path(bold_path).exists():
                pdfmetrics.registerFont(TTFont(reg_name, reg_path))
                pdfmetrics.registerFont(TTFont(bold_name, bold_path))
                if Path(ital_path).exists():
                    pdfmetrics.registerFont(TTFont(ital_name, ital_path))
                _FONT_FAMILY = reg_name
                _FONT_FAMILY_BOLD = bold_name
                _FONT_FAMILY_ITALIC = ital_name if Path(ital_path).exists() else bold_name
                break
        except Exception:
            continue
    _FONTS_INITIALIZED = True


class _NumberedCanvas(rl_canvas.Canvas if _REPORTLAB_AVAILABLE else object):
    """Two-pass canvas for dynamic 'Sayfa X / Y' page numbering and ASELSAN running header/footer."""
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._saved_page_states = []

    def showPage(self):
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        num_pages = len(self._saved_page_states)
        for state in self._saved_page_states:
            self.__dict__.update(state)
            self.draw_aselsan_decorations(num_pages)
            super().showPage()
        super().save()

    def draw_aselsan_decorations(self, total_pages: int):
        self.saveState()
        w, h = 595.27, 841.89  # A4 Portrait in points

        # Header (Pages > 1)
        if self._pageNumber > 1:
            self.setFillColor(rl_colors.HexColor("#002D62"))  # Aselsan Navy
            self.rect(36, h - 38, w - 72, 2, fill=True, stroke=False)
            self.setFont(_FONT_FAMILY_BOLD, 8)
            self.setFillColor(rl_colors.HexColor("#002D62"))
            self.drawString(36, h - 30, "ASELSAN A.Ş.  •  SAVUNMA VE GÜVENLİK SİSTEMLERİ")
            self.setFont(_FONT_FAMILY, 8)
            self.setFillColor(rl_colors.HexColor("#546E7A"))
            self.drawRightString(w - 36, h - 30, "CCTV Mühendislik & DORI Analiz Raporu")

        # Footer (All pages)
        self.setFillColor(rl_colors.HexColor("#CFD8DC"))
        self.rect(36, 42, w - 72, 0.75, fill=True, stroke=False)

        self.setFont(_FONT_FAMILY_BOLD, 7)
        self.setFillColor(rl_colors.HexColor("#002D62"))
        self.drawString(36, 30, "ASELSAN ELEKTRO-OPTİK & GÜVENLİK SİSTEMLERİ")

        self.setFont(_FONT_FAMILY, 7)
        self.setFillColor(rl_colors.HexColor("#B71C1C"))
        self.drawString(220, 30, "GİZLİLİK: KURUMSAL / HİZMETE ÖZEL")

        self.setFont(_FONT_FAMILY, 7)
        self.setFillColor(rl_colors.HexColor("#546E7A"))
        page_str = f"Sayfa {self._pageNumber} / {total_pages}"
        self.drawRightString(w - 36, 30, page_str)

        self.restoreState()


def export_pdf(
    path: str,
    cameras: List[Any],
    lens_mode: str,
    last_all_results: Dict[str, List[Any]],
    target_point: Any,
    selected_level_name: str,
    target_info_text: str,
    last_compliance_result: Optional[Dict[str, Any]] = None,
):
    """
    ASELSAN Kurumsal Kimliğine ve Savunma Sanayii Mühendislik Standartlarına tam uyumlu
    profesyonel PDF Raporu üretir. ReportLab mevcut değilse yedek motora geçer.
    """
    if not _REPORTLAB_AVAILABLE:
        # Fallback to simple PDF generator
        _export_pdf_fallback(path, cameras, lens_mode, last_all_results, target_point, selected_level_name, target_info_text)
        return

    _ensure_reportlab_fonts()

    doc = SimpleDocTemplate(
        path,
        pagesize=A4,
        leftMargin=36,
        rightMargin=36,
        topMargin=46,
        bottomMargin=54,
    )

    story = []
    page_w = 595.27 - 72  # Printable width = 523.27 pt

    # ASELSAN Corporate Color Palette
    c_navy = rl_colors.HexColor("#002D62")        # Primary Aselsan Blue
    c_steel = rl_colors.HexColor("#005A9C")       # Accent Blue
    c_light_bg = rl_colors.HexColor("#F4F6F9")    # Table header / card bg
    c_border = rl_colors.HexColor("#CFD8DC")      # Light gray border
    c_dark = rl_colors.HexColor("#212529")        # Body text
    c_success_bg = rl_colors.HexColor("#D1E7DD")
    c_success_fg = rl_colors.HexColor("#0F5132")
    c_warning_bg = rl_colors.HexColor("#FFF3CD")
    c_warning_fg = rl_colors.HexColor("#664D03")
    c_danger_bg = rl_colors.HexColor("#F8D7DA")
    c_danger_fg = rl_colors.HexColor("#842029")

    # Typography Styles
    styles = getSampleStyleSheet()

    style_title = ParagraphStyle(
        "AselsanTitle",
        fontName=_FONT_FAMILY_BOLD,
        fontSize=15,
        leading=18,
        textColor=rl_colors.white,
        alignment=0,
    )
    style_subtitle = ParagraphStyle(
        "AselsanSubtitle",
        fontName=_FONT_FAMILY_BOLD,
        fontSize=8.5,
        leading=11,
        textColor=rl_colors.HexColor("#90CAF9"),
        alignment=0,
    )
    style_sec_heading = ParagraphStyle(
        "AselsanSecHeading",
        fontName=_FONT_FAMILY_BOLD,
        fontSize=11,
        leading=14,
        textColor=c_navy,
        spaceBefore=10,
        spaceAfter=4,
    )
    style_body = ParagraphStyle(
        "AselsanBody",
        fontName=_FONT_FAMILY,
        fontSize=8.5,
        leading=11,
        textColor=c_dark,
    )
    style_body_bold = ParagraphStyle(
        "AselsanBodyBold",
        fontName=_FONT_FAMILY_BOLD,
        fontSize=8.5,
        leading=11,
        textColor=c_dark,
    )
    style_cell = ParagraphStyle(
        "AselsanCell",
        fontName=_FONT_FAMILY,
        fontSize=7.5,
        leading=9.5,
        textColor=c_dark,
    )
    style_cell_bold = ParagraphStyle(
        "AselsanCellBold",
        fontName=_FONT_FAMILY_BOLD,
        fontSize=7.5,
        leading=9.5,
        textColor=c_dark,
    )
    style_cell_header = ParagraphStyle(
        "AselsanCellHeader",
        fontName=_FONT_FAMILY_BOLD,
        fontSize=7.5,
        leading=9.5,
        textColor=rl_colors.white,
        alignment=1,
    )

    # ── 1. ASELSAN CORPORATE BANNER HEADER ──
    now_str = datetime.now().strftime("%d.%m.%Y %H:%M")
    doc_no = f"ASELSAN-CCTV-ENG-{datetime.now().strftime('%Y%m%d')}-01"

    banner_data = [
        [
            Paragraph("ASELSAN A.Ş.  •  SAVUNMA VE GÜVENLİK TEKNOLOJİLERİ SEKTÖR BAŞKANLIĞI", style_subtitle),
            Paragraph(f"DOKÜMAN NO: {doc_no}", ParagraphStyle("DocNo", fontName=_FONT_FAMILY_BOLD, fontSize=7.5, textColor=rl_colors.HexColor("#90CAF9"), alignment=2))
        ],
        [
            Paragraph("CCTV KAMERA GÖRÜŞ ALANI, DORI VE ŞARTNAME UYUMLULUK MÜHENDİSLİK RAPORU", style_title),
            Paragraph(f"GİZLİLİK: KURUMSAL / HİZMETE ÖZEL<br/>TARİH: {now_str}", ParagraphStyle("BannerMeta", fontName=_FONT_FAMILY, fontSize=7.5, leading=10, textColor=rl_colors.white, alignment=2))
        ]
    ]
    banner_table = Table(banner_data, colWidths=[page_w * 0.72, page_w * 0.28])
    banner_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), c_navy),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ("LEFTPADDING", (0, 0), (-1, -1), 10),
        ("RIGHTPADDING", (0, 0), (-1, -1), 10),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    story.append(banner_table)
    story.append(Spacer(1, 8))

    # ── 2. EXECUTIVE SUMMARY & KPI CARDS ──
    story.append(Paragraph("1. YÖNETİCİ ÖZETİ VE PROJE PARAMETRELERİ", style_sec_heading))

    total_cams = len(cameras)
    active_res_count = sum(len(res_list) for res_list in last_all_results.values())
    mode_str = mode_label(lens_mode)
    target_status_str = f"{target_point.name} ({target_point.x_m:.1f}m) - {selected_level_name}" if target_point.active else "Belirtilmedi"

    kpi_data = [
        [
            Paragraph("Toplam Kamera Adedi", style_body_bold),
            Paragraph(f"<b>{total_cams} Adet</b>", style_body),
            Paragraph("Analiz Modu", style_body_bold),
            Paragraph(f"<b>{mode_str}</b>", style_body),
        ],
        [
            Paragraph("Hesaplanan Optik Konfigürasyon", style_body_bold),
            Paragraph(f"<b>{active_res_count} Konfigürasyon</b>", style_body),
            Paragraph("Kontrol / Hedef Noktası", style_body_bold),
            Paragraph(target_status_str, style_body),
        ]
    ]
    kpi_table = Table(kpi_data, colWidths=[page_w * 0.25, page_w * 0.25, page_w * 0.25, page_w * 0.25])
    kpi_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), c_light_bg),
        ("BOX", (0, 0), (-1, -1), 0.75, c_border),
        ("INNERGRID", (0, 0), (-1, -1), 0.5, c_border),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    story.append(kpi_table)
    story.append(Spacer(1, 8))

    # ── 3. CAMERA HARDWARE & OPTICS MATRIX ──
    story.append(Paragraph("2. KAMERA DONANIM VE OPTİK YAPILANDIRMA MATRİSİ", style_sec_heading))

    cam_headers = ["Kamera Adı", "Model / Ürün", "Sensör", "Çözünürlük", "Lens (mm)", "Direk (m)", "Tilt (°)", "HFOV / VFOV", "IR (m)"]
    cam_rows = [[Paragraph(h, style_cell_header) for h in cam_headers]]

    for cam in cameras:
        results = last_all_results.get(cam.name, [])
        hfov_str = f"{results[0].hfov_deg:.1f}° / {results[0].vfov_deg:.1f}°" if results else "-"
        cam_rows.append([
            Paragraph(f"<b>{cam.name}</b>", style_cell_bold),
            Paragraph(cam.model_name, style_cell),
            Paragraph(cam.sensor_name, style_cell),
            Paragraph(cam.resolution_name.split(" (")[0], style_cell),
            Paragraph(f"{cam.focal_min_mm:g} - {cam.focal_max_mm:g}", style_cell),
            Paragraph(f"{cam.pole_height_m:g}", style_cell),
            Paragraph(f"{cam.tilt_deg:g}°", style_cell),
            Paragraph(hfov_str, style_cell),
            Paragraph(f"{cam.ir_range_m:g}m", style_cell),
        ])

    cam_table = Table(cam_rows, colWidths=[page_w * 0.12, page_w * 0.18, page_w * 0.09, page_w * 0.14, page_w * 0.11, page_w * 0.08, page_w * 0.08, page_w * 0.12, page_w * 0.08])
    cam_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), c_navy),
        ("GRID", (0, 0), (-1, -1), 0.5, c_border),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [rl_colors.white, c_light_bg]),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("ALIGN", (0, 1), (1, -1), "LEFT"),
    ]))
    story.append(cam_table)
    story.append(Spacer(1, 8))

    # ── 4. EN 62676-4 DORI MISSION & RANGE ANALYSIS ──
    story.append(Paragraph("3. EN 62676-4 STANDARDI DORI GÖREV VE MENZİL ANALİZİ", style_sec_heading))

    dori_headers = ["Kamera", "Lens", "Görev / Standart", "Sınıf", "PPM", "Optik Menzil", "Etkin Zemin", "Kör Nokta", "Durum"]
    dori_rows = [[Paragraph(h, style_cell_header) for h in dori_headers]]

    for cam_name, results in last_all_results.items():
        for res in results:
            for row in sorted(res.rows, key=lambda item: item.ppm, reverse=True):
                ground_str = "-" if row.ground_distance_m <= 0 else f"{row.ground_distance_m:.1f} m"
                status_color = c_success_fg if "Aktif" in row.status or "Uyumlu" in row.status else (c_warning_fg if "limit" in row.status else c_danger_fg)
                dori_rows.append([
                    Paragraph(row.camera, style_cell),
                    Paragraph(row.mode, style_cell),
                    Paragraph(row.level, style_cell_bold),
                    Paragraph(row.level_type, style_cell),
                    Paragraph(f"{row.ppm:g}", style_cell_bold),
                    Paragraph(f"{row.optical_distance_m:.1f} m", style_cell),
                    Paragraph(ground_str, style_cell),
                    Paragraph(f"{res.dead_zone_m:.1f} m", style_cell),
                    Paragraph(f"<font color='{status_color.hexval()}'><b>{row.status}</b></font>", style_cell),
                ])

    dori_table = Table(dori_rows, colWidths=[page_w * 0.11, page_w * 0.08, page_w * 0.22, page_w * 0.10, page_w * 0.08, page_w * 0.11, page_w * 0.11, page_w * 0.09, page_w * 0.10])
    dori_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), c_steel),
        ("GRID", (0, 0), (-1, -1), 0.5, c_border),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [rl_colors.white, c_light_bg]),
        ("TOPPADDING", (0, 0), (-1, -1), 3.5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3.5),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("ALIGN", (2, 1), (2, -1), "LEFT"),
    ]))
    story.append(dori_table)
    story.append(Spacer(1, 8))

    # ── 5. COMPLIANCE MATRIX (IF AVAILABLE) ──
    if last_compliance_result and isinstance(last_compliance_result, dict):
        matrix = last_compliance_result.get("matrix", [])
        if matrix:
            story.append(Paragraph("4. ŞARTNAME UYUMLULUK MATRİSİ (COMPLIANCE MATRIX)", style_sec_heading))
            comp_headers = ["Profil", "İster Tanımı", "Model", "Durum", "Mühendislik Kanıtı / Değerlendirme"]
            comp_rows = [[Paragraph(h, style_cell_header) for h in comp_headers]]

            for m_row in matrix:
                status = str(m_row.get("status", "Uyumsuz"))
                st_color = c_success_fg if "Uyumlu" in status else (c_warning_fg if "Kısmi" in status else c_danger_fg)
                comp_rows.append([
                    Paragraph(str(m_row.get("profile_name", m_row.get("profile_id", ""))), style_cell),
                    Paragraph(str(m_row.get("requirement", "")), style_cell),
                    Paragraph(str(m_row.get("camera_model", "")), style_cell_bold),
                    Paragraph(f"<font color='{st_color.hexval()}'><b>{status}</b></font>", style_cell),
                    Paragraph(str(m_row.get("evidence", "")), style_cell),
                ])

            comp_table = Table(comp_rows, colWidths=[page_w * 0.15, page_w * 0.32, page_w * 0.16, page_w * 0.10, page_w * 0.27])
            comp_table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), c_navy),
                ("GRID", (0, 0), (-1, -1), 0.5, c_border),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [rl_colors.white, c_light_bg]),
                ("TOPPADDING", (0, 0), (-1, -1), 3.5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3.5),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]))
            story.append(comp_table)
            story.append(Spacer(1, 8))

    # ── 6. ENGINEERING RECOMMENDATIONS & DEAD ZONE WARNINGS ──
    story.append(Paragraph("5. OTOMATİK KURULUM VE MÜHENDİSLİK TAVSİYELERİ", style_sec_heading))

    recs = []
    for cam_name, results in last_all_results.items():
        for res in results:
            for rec in res.recommendations:
                recs.append(f"<b>[{cam_name} - {mode_label(res.mode)}]:</b> {rec}")

    if not recs:
        recs.append("Optik analiz kriterlerine göre tüm kameralar nominal çalışma parametreleri içerisindedir.")

    rec_data = [[Paragraph(f"• {r}", style_cell)] for r in recs]
    rec_table = Table(rec_data, colWidths=[page_w])
    rec_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), c_light_bg),
        ("BOX", (0, 0), (-1, -1), 0.75, c_navy),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
    ]))
    story.append(rec_table)

    # Build Document using NumberedCanvas
    doc.build(story, canvasmaker=_NumberedCanvas)


def _export_pdf_fallback(
    path: str,
    cameras: List[Any],
    lens_mode: str,
    last_all_results: Dict[str, List[Any]],
    target_point: Any,
    selected_level_name: str,
    target_info_text: str
):
    """Fallback simple text PDF generator if reportlab is unavailable."""
    lines = [
        "ASELSAN CCTV ANALİZ VE UYUMLULUK RAPORU",
        f"Tarih: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        f"Kamera Sayisi: {len(cameras)}",
        f"Analiz Modu: {mode_label(lens_mode)}",
        "",
    ]
    for camera in cameras:
        lines.extend([
            f"Kamera: {camera.name}",
            f"Sensor: {camera.sensor_name} | Cozunurluk: {camera.resolution_name}",
            f"Lens: {camera.focal_min_mm:g}-{camera.focal_max_mm:g} mm | Tilt: {camera.tilt_deg:g} derece",
            f"Konum: X {camera.pos_x_m:g} m, Y {camera.pos_y_m:g} m, Yon {camera.heading_deg:g} derece",
            f"Model: {camera.model_name} | IR: {camera.ir_range_m:g} m | Min lux: {camera.min_lux:g}",
        ])
        for result in last_all_results.get(camera.name, []):
            lines.append(
                f"  {mode_label(result.mode)}: HFOV {result.hfov_deg:.1f}, VFOV {result.vfov_deg:.1f}, "
                f"Kor nokta {result.dead_zone_m:.1f} m"
            )
            for row in sorted(result.rows, key=lambda item: item.ppm, reverse=True):
                ground = "-" if row.ground_distance_m <= 0 else f"{row.ground_distance_m:.2f} m"
                lines.append(f"    {row.level:<22} {row.ppm:g} px/m  opt {row.optical_distance_m:.2f} m  etkin {ground}  {row.status}")
            for recommendation in result.recommendations:
                lines.append(f"    Oneri: {recommendation}")
        lines.append("")
    if target_point.active:
        lines.extend([
            "Kontrol Noktasi",
            f"Ad: {target_point.name}",
            f"Konum: X {target_point.x_m:g} m, Y {target_point.y_m:g} m",
            f"Gereken Seviye: {selected_level_name}",
            target_info_text,
            "",
        ])
    write_simple_pdf(path, "ASELSAN CCTV ANALİZ VE UYUMLULUK RAPORU", lines)



def export_excel(
    path: str,
    cameras: List[Any],
    lens_mode: str,
    last_all_results: Dict[str, List[Any]],
    target_point: Any,
    selected_level_name: str,
    last_compliance_result: Optional[Dict[str, Any]] = None
):
    """Gelişmiş biçimlendirilmiş çok sekmeli Excel (.xlsx) raporu üretir."""
    if openpyxl is None:
        raise RuntimeError("Excel aktarımı için 'openpyxl' kütüphanesi gerekli.")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")

    title_font = Font(name="Segoe UI", size=13, bold=True, color="1F4E78")
    bold_font = Font(name="Segoe UI", size=10, bold=True)
    normal_font = Font(name="Segoe UI", size=10)

    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")

    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9")
    )

    status_styles = {
        "Aktif": (PatternFill("solid", fgColor="E2EFDA"), Font(name="Segoe UI", size=10, bold=True, color="375623")),
        "Uyumlu": (PatternFill("solid", fgColor="E2EFDA"), Font(name="Segoe UI", size=10, bold=True, color="375623")),
        "Kör noktada": (PatternFill("solid", fgColor="FCE4D6"), Font(name="Segoe UI", size=10, bold=True, color="C65911")),
        "Uyumsuz": (PatternFill("solid", fgColor="FCE4D6"), Font(name="Segoe UI", size=10, bold=True, color="C65911")),
        "Geometrik limit": (PatternFill("solid", fgColor="FFF2CC"), Font(name="Segoe UI", size=10, bold=True, color="806000")),
        "Kısmi": (PatternFill("solid", fgColor="FFF2CC"), Font(name="Segoe UI", size=10, bold=True, color="806000")),
        "Bulunamadı": (PatternFill("solid", fgColor="EDEDED"), Font(name="Segoe UI", size=10, italic=True, color="595959")),
    }

    # -------------------------------------------------------------
    # SHEET 1: Kamera Analizi
    # -------------------------------------------------------------
    ws1 = wb.create_sheet(title="Kamera Analizi")
    ws1.views.sheetView[0].showGridLines = True

    ws1.cell(row=1, column=1, value="CCTV Kamera Görüş Alanı ve Kapsama Analizi Tablosu").font = title_font
    ws1.cell(row=2, column=1, value=f"Rapor Tarihi: {datetime.now().strftime('%Y-%m-%d %H:%M')} | Toplam Kamera: {len(cameras)} | Mod: {mode_label(lens_mode)}").font = normal_font

    headers1 = [
        "Kamera", "Model", "Lens", "Odak (mm)", "HFOV (°)", "VFOV (°)",
        "Kör Nokta (m)", "Kör Nokta (m²)", "Görev / Algoritma", "Sınıf", "PPM",
        "Optik Menzil (m)", "Etkin Menzil (m)", "Durum"
    ]
    start_row = 4
    for col_idx, h in enumerate(headers1, 1):
        cell = ws1.cell(row=start_row, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    current_row = start_row + 1
    for results in last_all_results.values():
        for result in results:
            for row in sorted(result.rows, key=lambda item: item.ppm, reverse=True):
                r_vals = [
                    row.camera,
                    result.camera.model_name,
                    row.mode,
                    result.focal_mm,
                    round(result.hfov_deg, 1),
                    round(result.vfov_deg, 1),
                    round(result.dead_zone_m, 1),
                    round(result.dead_zone_area_m2, 1),
                    row.level,
                    row.level_type,
                    row.ppm,
                    round(row.optical_distance_m, 2),
                    None if row.ground_distance_m <= 0 else round(row.ground_distance_m, 2),
                    row.status
                ]
                for col_idx, val in enumerate(r_vals, 1):
                    c = ws1.cell(row=current_row, column=col_idx, value=val)
                    c.font = normal_font
                    c.border = thin_border
                    if col_idx in (4, 5, 6, 7, 8, 11, 12, 13):
                        c.alignment = right_align
                    elif col_idx in (3, 10, 14):
                        c.alignment = center_align
                    else:
                        c.alignment = left_align

                    if col_idx == 14 and row.status in status_styles:
                        fill, font = status_styles[row.status]
                        c.fill = fill
                        c.font = font
                current_row += 1

    # -------------------------------------------------------------
    # SHEET 2: Compliance Matrix (Eğer varsa)
    # -------------------------------------------------------------
    if last_compliance_result and last_compliance_result.get("matrix"):
        ws2 = wb.create_sheet(title="Compliance Matrix")
        ws2.views.sheetView[0].showGridLines = True
        ws2.cell(row=1, column=1, value="Teknik Şartname Uyum Matrisi (Compliance Matrix)").font = title_font

        headers2 = ["Profil", "İster", "Kamera / Model", "Durum", "Kanıt / Not"]
        for col_idx, h in enumerate(headers2, 1):
            cell = ws2.cell(row=3, column=col_idx, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align

        c_row = 4
        for r in last_compliance_result.get("matrix", []):
            st = r.get("status", "")
            vals = [
                r.get("profile_name", ""),
                r.get("requirement", ""),
                r.get("camera_model", ""),
                st,
                r.get("evidence", "")
            ]
            for col_idx, val in enumerate(vals, 1):
                c = ws2.cell(row=c_row, column=col_idx, value=val)
                c.font = normal_font
                c.border = thin_border
                c.alignment = center_align if col_idx == 4 else left_align
                if col_idx == 4 and st in status_styles:
                    fill, font = status_styles[st]
                    c.fill = fill
                    c.font = font
            c_row += 1

    # -------------------------------------------------------------
    # SHEET 3: Proje Özeti
    # -------------------------------------------------------------
    ws3 = wb.create_sheet(title="Proje Özeti")
    ws3.views.sheetView[0].showGridLines = True
    ws3.cell(row=1, column=1, value="CCTV Proje ve Donanım Metraj Özeti").font = title_font

    summary_data = [
        ("Proje Tarihi", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Toplam Kamera Sayısı", len(cameras)),
        ("Analiz Modu", mode_label(lens_mode)),
        ("Kontrol Noktası Durumu", "Aktif" if target_point.active else "Pasif"),
        ("Kontrol Noktası Adı", target_point.name if target_point.active else "-"),
        ("Gereken Seviye", selected_level_name),
    ]

    for idx, (k, v) in enumerate(summary_data, 3):
        ck = ws3.cell(row=idx, column=1, value=k)
        ck.font = bold_font
        ck.border = thin_border
        cv = ws3.cell(row=idx, column=2, value=v)
        cv.font = normal_font
        cv.border = thin_border

    ws3.cell(row=11, column=1, value="Kamera Donanım Listesi").font = title_font
    cam_headers = ["Kamera Adı", "Model", "Sensör", "Çözünürlük", "Lens (mm)", "Tilt (°)", "Montaj Yüksekliği (m)", "IR (m)", "Min Lux", "Konum (X, Y)"]
    for col_idx, h in enumerate(cam_headers, 1):
        cell = ws3.cell(row=13, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    crow = 14
    for cam in cameras:
        cvals = [
            cam.name,
            cam.model_name,
            cam.sensor_name,
            cam.resolution_name,
            f"{cam.focal_min_mm:g}-{cam.focal_max_mm:g}",
            cam.tilt_deg,
            cam.pole_height_m,
            cam.ir_range_m,
            cam.min_lux,
            f"X: {cam.pos_x_m:g}, Y: {cam.pos_y_m:g}"
        ]
        for col_idx, val in enumerate(cvals, 1):
            c = ws3.cell(row=crow, column=col_idx, value=val)
            c.font = normal_font
            c.border = thin_border
            c.alignment = left_align
        crow += 1

    # Sütun genişliklerini otomatik ayarla
    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    lines = str(cell.value).split("\n")
                    for l in lines:
                        max_len = max(max_len, len(l))
            sheet.column_dimensions[col_letter].width = max(max_len + 4, 12)

    wb.save(path)


def export_compliance_excel(path: str, last_compliance_result: Dict[str, Any]):
    """Sadece Compliance Matrix için özel Excel (.xlsx) raporu üretir."""
    if openpyxl is None:
        raise RuntimeError("Excel aktarımı için 'openpyxl' kütüphanesi gerekli.")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Compliance Matrix"
    ws.views.sheetView[0].showGridLines = True

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    title_font = Font(name="Segoe UI", size=13, bold=True, color="1F4E78")
    normal_font = Font(name="Segoe UI", size=10)
    bold_font = Font(name="Segoe UI", size=10, bold=True)

    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")

    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9")
    )

    status_styles = {
        "Uyumlu": (PatternFill("solid", fgColor="E2EFDA"), Font(name="Segoe UI", size=10, bold=True, color="375623")),
        "Uyumsuz": (PatternFill("solid", fgColor="FCE4D6"), Font(name="Segoe UI", size=10, bold=True, color="C65911")),
        "Kısmi": (PatternFill("solid", fgColor="FFF2CC"), Font(name="Segoe UI", size=10, bold=True, color="806000")),
        "Bulunamadı": (PatternFill("solid", fgColor="EDEDED"), Font(name="Segoe UI", size=10, italic=True, color="595959")),
    }

    ws.cell(row=1, column=1, value="Teknik Şartname Uyum Matrisi (Compliance Matrix)").font = title_font
    ws.cell(row=2, column=1, value=f"Tarih: {datetime.now().strftime('%Y-%m-%d %H:%M')}").font = normal_font

    headers = ["Profil", "İster", "Kamera / Model", "Durum", "Kanıt / Not"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    c_row = 5
    for r in last_compliance_result.get("matrix", []):
        st = r.get("status", "")
        vals = [
            r.get("profile_name", ""),
            r.get("requirement", ""),
            r.get("camera_model", ""),
            st,
            r.get("evidence", "")
        ]
        for col_idx, val in enumerate(vals, 1):
            c = ws.cell(row=c_row, column=col_idx, value=val)
            c.font = normal_font
            c.border = thin_border
            c.alignment = center_align if col_idx == 4 else left_align
            if col_idx == 4 and st in status_styles:
                fill, font = status_styles[st]
                c.fill = fill
                c.font = font
        c_row += 1

    # Skor ve Öneri Bölümü
    c_row += 2
    ws.cell(row=c_row, column=1, value="Öneri ve Skor Özeti").font = title_font
    c_row += 1
    rec_cell = ws.cell(row=c_row, column=1, value=f"Öneri: {last_compliance_result.get('recommendation', '')}")
    rec_cell.font = bold_font
    c_row += 2

    for score_item in last_compliance_result.get("camera_scores", [])[:10]:
        t = f"- {score_item.get('profile_name', '')}: {score_item.get('camera_model', '')} -> Skor: {score_item.get('score', 0)} ({score_item.get('verdict', '')})"
        ws.cell(row=c_row, column=1, value=t).font = normal_font
        c_row += 1

    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                lines = str(cell.value).split("\n")
                for l in lines:
                    max_len = max(max_len, len(l))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    wb.save(path)
